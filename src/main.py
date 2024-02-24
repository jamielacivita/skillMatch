import DataPool
import pickle
import openpyxl
import Host as Host
import Extern as Extern
import Skills as Skills
import Match as Match

import logging.config
#logging.config.fileConfig("/home/jamie/source/python/skillMatch/src/logging_config.ini")
logging.config.fileConfig("/home/jamie/PycharmProjects/skillMatch/src/logging_config.ini")
log = logging.getLogger(__name__)

def get_match_object_lst():
    host_dataframe = get_host_dataframe()
    host_objects_list = parse_host_dataframe(host_dataframe)

    extern_dataframe = get_extern_dataframe()
    extern_objects_list = parse_extern_dataframe(extern_dataframe)

    match_obj_list = []
    for extern in extern_objects_list:
        for host in host_objects_list:
            # make a empty match object.
            match_obj = Match.Match(extern,host)
            match_obj_list.append(match_obj)

    return match_obj_list

def host_generate_field_headings(filename="/home/jamie/PycharmProjects/skillMatch/data/240218-055238_Host.xlsx"):
    """
    Given a path to a host file return a list of headings for the records.
    :return:
    """
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active
    for row in range(0, dataframe1.max_row):
    #for row in range(1, 2):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            cell = col[row]
            if cell.column_letter == "AJ":
                print(f" {cell.column_letter} : {cell.value}")

def get_host_dataframe(filename="/home/jamie/PycharmProjects/skillMatch/data/240218-055238_Host.xlsx"):
    """
    Given a path to a host file return a dataframe records.
    :return:
    """
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active

    return dataframe1

def parse_host_dataframe(dataframe):
    """
    :param dataframe:
    :return: a list of lists holding the information in the dataframe.
    """

    out_lst = []

    FIRST_DATA_ROW = 2


    for row in dataframe.iter_rows(FIRST_DATA_ROW, dataframe.max_row):
    ## each row represents a host
    ## create and populate host object
        host_obj = Host.Host(row)
        out_lst.append(host_obj)

    return out_lst


def get_extern_dataframe():
    print("foo_populate_Extern")
    filename = r"/home/jamie/PycharmProjects/skillMatch/data/240218-055306_Educator.xlsx"

    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active

    return dataframe1

def parse_extern_dataframe(dataframe):
    out_lst = []
    FIRST_DATA_ROW = 2
    for row in dataframe.iter_rows(FIRST_DATA_ROW, dataframe.max_row):
        extern_obj = Extern.Extern(row)
        out_lst.append(extern_obj)

    return out_lst



if __name__ == "__main__":
    match_object_list = get_match_object_lst()
    for match in match_object_list:
        print(match)

    print("Done.")



