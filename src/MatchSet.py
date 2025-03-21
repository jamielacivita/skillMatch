import openpyxl
import csv
import time
import sys

import Host as Host
import Extern as Extern
import Match as Match
from prettytable import PrettyTable as pt
import data_config

import logging.config
# logging.config.fileConfig("/home/jamie/source/python/skillMatch/src/logging_config.ini")
#logging.config.fileConfig("/home/jamie/Source/Python/skillMatch/src/logging_config.ini")
logging.config.fileConfig(data_config.LoggingConfigurationFile)

log = logging.getLogger(__name__)


class MatchSet():

    def __init__(self):
        log.debug("1. Generating a Matchset Object.")
        self.extern_obj_lst = self.get_extern_obj_lst()
        log.debug(f"length of extern_obj_list : {len(self.extern_obj_lst)}")
        self.host_obj_lst = self.get_host_objects_lst()
        log.debug(f"length of host_obj_list : {len(self.host_obj_lst)}")
        self.match_obj_lst = self.get_match_object_lst()
        log.debug(f"length of match_obj_list : {len(self.match_obj_lst)}")

    def parse_host_dataframe(self, dataframe=None, dataframecsv=None):
        """
        :param dataframe:
        :return: a list of lists holding the information in the dataframe.
        """
        if dataframe is not None:
            out_lst = []

            FIRST_DATA_ROW = 2

            for row in dataframe.iter_rows(FIRST_DATA_ROW, dataframe.max_row):
                ## each row represents a host
                ## create and populate host object
                host_obj = Host.Host(row)
                out_lst.append(host_obj)

            return out_lst

        elif dataframecsv is not None:
            out_lst = []
            # using a csvdataframe
            # log.debug("using a csv dataframe.")
            FIRST_DATA_ROW = 1
            LAST_DATA_ROW = len(dataframecsv)

            for row in dataframecsv[FIRST_DATA_ROW:LAST_DATA_ROW]:
                host_obj = Host.Host(row)
                out_lst.append(host_obj)

            return out_lst

    def get_extern_dataframe(self):
        """
        requires a path to the xlsx file with the educator (teacher) information.
        :return: an openpyxl.worksheet.worksheet.Worksheet object.
        """
        filename = data_config.EdApplicationPath
        if (".xlsx" in filename):
            # log.debug("This is an XLSX file.")
            dataframe = openpyxl.load_workbook(filename)
            dataframe1 = dataframe.active

        elif (".csv" in filename):
            #log.debug("This is a CSV file.")
            dataframe1 = []
            with open(filename, encoding='latin-1') as csvfile:
                spamreader = csv.reader(csvfile)
                for row in spamreader:
                    dataframe1.append(row)

        return dataframe1

    def parse_extern_dataframe(self, dataframe=None, dataframecsv=None):
        """
        Given a set of data return a list of extern objects.
        """
        out_lst = []

        if dataframe is not None:
            # using a xlsxdataframe
            # log.debug("using an xlsx dataframe.")
            FIRST_DATA_ROW = 2
            for row in dataframe.iter_rows(FIRST_DATA_ROW, dataframe.max_row):
                extern_obj = Extern.Extern(row)
                out_lst.append(extern_obj)


        elif dataframecsv is not None:
            # using a csvdataframe
            # log.debug("using a csv dataframe.")
            FIRST_DATA_ROW = 1
            LAST_DATA_ROW = len(dataframecsv)

            for row in dataframecsv[FIRST_DATA_ROW:LAST_DATA_ROW]:
                extern_obj = Extern.Extern(row)
                out_lst.append(extern_obj)

        return out_lst

    def get_host_objects_lst(self):
        host_dataframe = self.get_host_dataframe()
        host_objects_lst = self.parse_host_dataframe(dataframecsv=host_dataframe)
        return host_objects_lst

    def get_extern_obj_lst(self):
        """
        Given a collection of extern data (extern dataframe)
        :return: a list of extern objects.
        """
        extern_dataframe = self.get_extern_dataframe()
        extern_objects_lst = self.parse_extern_dataframe(dataframecsv=extern_dataframe)
        return extern_objects_lst

    def get_match_object_lst(self):
        """
        Input a list of extern objects and a list of host objects.
        :return: a list of match objects.
        """
        match_obj_list = []
        for extern in self.get_extern_obj_lst():
            for host in self.get_host_objects_lst():
                match_obj = Match.Match(extern, host)
                match_obj_list.append(match_obj)

        return match_obj_list

    def get_host_dataframe(self, host_filename="/home/jamie/PycharmProjects/skillMatch/data/240314-025428_HostApplication_Final.xlsx"):
        """
        Given a path to a host file return a dataframe records.
        :return:
        """

        host_filename = data_config.HostApplicationPath
        if (".xlsx" in host_filename):
            log.debug("Using a .xlsx file for the hosts data.")
            dataframe = openpyxl.load_workbook(host_filename)
            dataframe1 = dataframe.active

            return dataframe1

        elif (".csv" in host_filename):
            #log.debug("Using a .csv file for the hosts data.")
            dataframe1 = []

            with open(host_filename, encoding='latin-1') as csvfile:
                spamreader = csv.reader(csvfile)
                for row in spamreader:
                    dataframe1.append(row)

            return dataframe1

    def get_number_hosts(self):
        return len(self.host_obj_lst)

    def get_number_externs(self):
        return len(self.extern_obj_lst)

    def get_number_matches(self):
        return len(self.match_obj_lst)

    def get_match_chart_header_row(self):
        header_row = []
        header_row.append("Extern ID")                          #match.get_extern_id()
        header_row.append("Host ID")                            #match.get_host_id()
        header_row.append("EXTERN")                             #match.get_extern_name()
        header_row.append("EXTERN CITY")                        #atch.get_extern_city()
        header_row.append("HOST")                               #match.get_host_name()
        header_row.append("HOST CITY")                          #match.get_host_city()
        header_row.append("HOST NO EXP NEEDED")                 #match.get_host_no_skills_needed()
        header_row.append("DISTANCE")                           #match.get_distance()
        header_row.append("DISTANCE NOTES")                     #match.get_distance_notes()
        header_row.append("REMOTE MATCH")                       #match.get_remote_match()
        header_row.append("REMOTE ONLY")                        #match.get_extern_remote_only()
        header_row.append("EXTERN LOCATION OPENNESS")           #match.get_extern_open_to()
        header_row.append("HOST LOCATION OPENNESS")             #match.get_host_work_done_remotely()
        header_row.append("STEM experience match")              #match.get_stem_experience_match_score()
        header_row.append("STEM experience matched on")         #match.get_stem_experience_match_notes()
        header_row.append("Biz Skills Score")                   #match.get_biz_skills_match_score()
        header_row.append("Biz Skills Match")                   #match.get_biz_skills_match_notes()
        header_row.append("Work Style Score")                   #match.get_work_style_match_score()
        header_row.append("Work Style Notes")                   #match.get_work_style_match_notes()
        header_row.append("Teaching Needs Match")               #match.get_teaching_needs_match()
        header_row.append("Curriculum Needs Match")             #match.get_curriculum_design_match()
        header_row.append("Total Experience Match")             #match.get_total_score()
        header_row.append("Min Interns")                        #match.get_min_externs()
        header_row.append("Max Interns")                        #match.get_max_externs()
        header_row.append("(first) Location")                   #match.get_arson()
        header_row.append("(second) Location")                  #match.get_basin()
        header_row.append("Teaching Openness")                  #match.get_carrear()
        header_row.append("Business Skills")                    #match.get_defend()
        header_row.append("STEM Domains")                       #match.get_endorse()
        header_row.append("STEM Interest")                      #match.get_fortify()
        header_row.append("Hours")                              #match.get_gavel()

        return header_row

    def get_match_chart_data_rows(self, extern_id = None):
        out_rows = []
        for match in self.match_obj_lst:
            out_row = []

            out_row.append(f"{match.get_extern_id()}")
            out_row.append(f"{match.get_host_id()}")
            out_row.append(f"{match.get_extern_name()}")
            out_row.append(f"{match.get_extern_city()}")
            out_row.append(f"{match.get_host_name()}")
            out_row.append(f"{match.get_host_city()}")
            out_row.append(f"{match.get_host_no_skills_needed()}")
            out_row.append(f"{match.get_distance()}")
            out_row.append(f"{match.get_distance_notes()}")
            out_row.append(f"{match.get_remote_match()}")
            out_row.append(f"{match.get_extern_remote_only()}")
            out_row.append(f"{match.get_extern_open_to()}")
            out_row.append(f"{match.get_host_work_done_remotely()}")
            out_row.append(f"{match.get_stem_experience_match_score()}")
            out_row.append(f"{match.get_stem_experience_match_notes()}")
            out_row.append(f"{match.get_biz_skills_match_score()}")
            out_row.append(f"{match.get_biz_skills_match_notes()}")
            out_row.append(f"{match.get_work_style_match_score()}")
            out_row.append(f"{match.get_work_style_match_notes()}")
            out_row.append(f"{match.get_teaching_needs_match()}")
            out_row.append(f"{match.get_curriculum_design_match()}")
            out_row.append(f"{match.get_total_score()}")
            out_row.append(f"{match.get_min_externs()}")
            out_row.append(f"{match.get_max_externs()}")
            out_row.append(f"{match.get_arson()}")
            out_row.append(f"{match.get_basin()}")
            out_row.append(f"{match.get_carrear()}")
            out_row.append(f"{match.get_defend()}")
            out_row.append(f"{match.get_endorse()}")
            out_row.append(f"{match.get_fortify()}")
            out_row.append(f"{match.get_gavel()}")

            out_rows.append(out_row)

        return out_rows

    def print_match_chart(self):

        out_table = pt()
        out_table.field_names = self.get_match_chart_header_row()
        for row in self.get_match_chart_data_rows():
            out_table.add_row(row)

        print(out_table)
        return None

    def print_extern_city(self):
        """
        :return: None : printed output is a list of cities that Externs live in.
        """
        for extern_obj in self.extern_obj_lst:
            print(extern_obj.get_city_you_live_in())

    def save_match_chart_csv(self):
        """
        :return: None - effect is to save the match chart data to a CSV.
        """
        #path = f"/home/jamie/PycharmProjects/skillMatch/data/"
        #path = f"/home/jamie/Source/Python/skillMatch/data/Output/chart/"
        path = f"/home/jamie/source/skillMatch/data/Output/chart/"
        filename = f"{time.strftime('%Y%m%d-%H%M%S')}_out.csv"
        path_to_file = path + filename
        with open(path_to_file, "w", newline='') as csvfile:
            match_chart_write = csv.writer(csvfile)
            match_chart_write.writerow(self.get_match_chart_header_row())
            for row in self.get_match_chart_data_rows():
                match_chart_write.writerow(row)
        csvfile.close()
        return None

    def print_host_dossiers(self):
        for host in self.get_host_objects_lst():
            host.save_dossier()

    def print_extern_dossiers(self):

        for extern in self.get_extern_obj_lst():
            extern.save_dossier()

